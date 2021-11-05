from Source.classroom import Classroom
from Source.lecture import Lecture
from Source.lecturer import Lecturer
from Source.meeting import Meeting
from Source.course import Course
from Source.semester import Semester
import os
import json
import openpyxl
import csv
from prettytable import PrettyTable,ALL
import pandas as pd
import pickle
from Source.time_ import convert2timegap
import termcolor

class Timetable:
    available_semesters=[i for i in range(1,9)]
    dit_schedule=os.path.join('','dit_factors','full_dit_info.xlsx')
    dit_classrooms=os.path.join('','dit_factors','dit_classrooms.xlsx')
    dit_teachers=os.path.join('','dit_factors','teachers.csv')
    dit_courses=os.path.join('','dit_factors','dit_courses.xlsx')
    dit_courses_old=os.path.join('','dit_factors','dit_courses_old.xlsx')
    dit_courses_extra=os.path.join('','dit_factors','dit_courses_extra.xlsx')
    
    def __init__(self,id):
        self.id=id
        self.classrooms=list()
        self.lecturers=list()
        self.lectures=list()
        self.meetings=list()
        self.courses=list()
        self.semester_info=list()
        self.ignore_info=list()


    def import_full_dit_schedule(self):
        # Read classrooms
        wb_obj=openpyxl.load_workbook(Timetable.dit_classrooms)
        sheet_obj=wb_obj.active
        total_rows=sheet_obj.max_row+1
        total_columns=sheet_obj.max_column+1
        for i in range(1,total_rows):
            row=list()
            for j in range(1,total_columns):
                cell_obj=sheet_obj.cell(row=i,column=j)
                row.append(cell_obj.value)
            if i!=1:
                self.classrooms.append(Classroom(row[0],row[1],int(row[2])))
        del wb_obj
        del sheet_obj

        # Read Lectures
        RF=None
        with open(Timetable.dit_teachers,'r') as f:
            RF=csv.reader(f)
            start=True
            for name,rank,mail,id in RF:
                if start:
                    start=False
                    continue
                self.lecturers.append(Lecturer(id,name,mail,rank))

        #Read Courses
        wb_obj=openpyxl.load_workbook(Timetable.dit_courses)
        sheet_obj=wb_obj.active
        total_rows=sheet_obj.max_row+1
        total_columns=sheet_obj.max_column+1
        start=True
        for i in range(1,total_rows):
            if start:
                start=False
                continue
            row=list()
            for j in range(1,total_columns):
                cell_obj=sheet_obj.cell(i,j)
                row.append(cell_obj.value)
            self.courses.append(Course(row[0],row[1],int(row[2]),row[3],int(row[4]),int(row[5]),int(row[6]),int(row[7])))
        del wb_obj
        del sheet_obj

        wb_obj=openpyxl.load_workbook(Timetable.dit_courses_old)
        sheet_obj=wb_obj.active
        total_rows=sheet_obj.max_row+1
        total_columns=sheet_obj.max_column+1
        start=True
        for i in range(1,total_rows):
            if start:
                start=False
                continue
            row=list()
            for j in range(1,total_columns):
                cell_obj=sheet_obj.cell(i,j)
                row.append(cell_obj.value)
            self.courses.append(Course(row[0],row[1],int(row[2]),row[3],int(row[4]),int(row[5]),int(row[6]),int(row[7])))

        del wb_obj
        del sheet_obj

        #Read Extra lessons
        wb_obj=openpyxl.load_workbook(Timetable.dit_courses_extra)
        sheet_obj=wb_obj.active
        total_rows=sheet_obj.max_row+1
        total_columns=sheet_obj.max_column+1
        start=True

        for i in range(1,total_rows):
            if start:
                start=False
                continue
            row=list()
            for j in range(1,total_columns):
                cell_obj=sheet_obj.cell(i,j)
                row.append(cell_obj.value)
            self.courses.append(Course(row[0],row[1],int(row[2]),row[3],int(row[4]),int(row[5]),int(row[6]),int(row[7])))

        del wb_obj
        del sheet_obj
        # Read Meetings
        wb_obj=openpyxl.load_workbook(Timetable.dit_schedule)
        sheet_obj=wb_obj.active
        total_rows=sheet_obj.max_row+1
        total_columns=sheet_obj.max_column+1
        autoincreament_id=1
        for i in range(1,total_rows):
            row=list()
            for j in range(1,total_columns):
                cell_obj=sheet_obj.cell(row=i,column=j)
                row.append(cell_obj.value)
            if i==1: continue
            hour_presentation=row[1]
            data=hour_presentation.split('-')
            start_hour=data[0].split(':')[0]
            end_hour=data[1].split(':')[0]
            duration=int(end_hour)-int(start_hour)
            lecture_instance=Lecture(row[3].strip(),duration,self.classrooms[self.classrooms.index(row[4].strip())],self.lecturers[self.lecturers.index(row[5])])
            self.lectures.append(lecture_instance)
            self.meetings.append(Meeting(autoincreament_id,data[0],data[1],row[0],self.courses[self.courses.index(row[2].strip()+"-"+str(row[6]))],int(row[6]),lecture_instance))
            autoincreament_id+=1

        self.create_schema()
        
    def create_schema(self):
        # Split into semester and lecturers spliting
        self.semester_info=list()
        for meeting in self.meetings:
             if meeting.course.get_semester() in self.semester_info: 
                continue
             self.semester_info.append(Semester(meeting.course.get_semester()))

        for meeting in self.meetings:
            self.semester_info[self.semester_info.index(meeting.course.get_semester())].add_meeting(meeting)
            period_zones=meeting.timezone()
            for zone in period_zones:
                self.lecturers[self.lecturers.index(meeting.lecture.lecturer.identifier)].add_job(meeting.day,meeting.course.title,zone)

        for semester in self.semester_info:
            if "-" in semester.id:
                id_semester=semester.id.split("-")
                csemester=int(id_semester[0])
                cflow=id_semester[1]
                for course in self.courses:
                    if course.semester==csemester and course.flow==cflow:
                        semester.add_course(course)
            else:
                csemester=int(semester.id)
                for course in self.courses:
                    if course.semester==csemester:
                        semester.add_course(course)



    def print_semester_program(self,semester):
        table=PrettyTable()
        table.hrules=ALL
        namefield=[' ']
        namefield.extend(Semester.days)
        table.field_names=namefield
        widths={name:30 for name in namefield}
        table._max_width=widths
        for timestamp in Semester.timezones:
            row=list()
            start_time=int(timestamp.split('-')[0].split(':')[0])
            end_time=int(timestamp.split('-')[1].split(':')[0])
            row.append(timestamp)
            for day in Semester.days:
                meeting_schedule=str()
                for meeting in self.semester_info[self.semester_info.index(semester)].meetings[day]:
                        meeting_start_time=int(meeting.start_hour.split(":")[0])
                        meeting_end_time=int(meeting.end_hour.split(":")[0])
                        if meeting_start_time<=start_time and meeting_end_time>=end_time:
                            meeting_schedule+=meeting.description()+"\n"
                row.append(meeting_schedule)
            table.add_row(row)
        print(table)
    
    def export_semester_program(self,semester):
        header=['  ']
        header.extend(Semester.days)
        semester_schedule={h:list() for h in header}
        for timestamp in Semester.timezones:
            row=list()
            start_time=int(timestamp.split('-')[0].split(':')[0])
            end_time=int(timestamp.split('-')[1].split(':')[0])
            row.append(timestamp)
            for day in Semester.days:
                meeting_schedule=""
                for meeting in self.semester_info[self.semester_info.index(semester)].meetings[day]:
                    meeting_start_time=int(meeting.start_hour.split(':')[0])
                    meeting_end_time=int(meeting.end_hour.split(':')[0])
                    if meeting_start_time<=start_time and meeting_end_time>=end_time:
                        meeting_schedule+=meeting.description()+"\n"
                row.append(meeting_schedule)
            i=0
            for h in header:
                semester_schedule[h].append(row[i])
                i+=1
        name_of_pickle_file="dit_"+str(semester)+".pkl"
        name_of_xlsx_file="dit_"+str(semester)+".xlsx"
        path_to_serialize=os.path.join('','Schedule_Serialization')
        path_to_xlsx=os.path.join('','Schedule_Xlsx')
        with open(os.path.join(path_to_serialize,name_of_pickle_file),'wb') as WF:
            pickle.dump(semester_schedule,WF)
        df=pd.DataFrame(semester_schedule)
        excelwriter=pd.ExcelWriter(os.path.join(path_to_xlsx,name_of_xlsx_file))
        df.to_excel(excelwriter)
        excelwriter.save()

    def validate_lessons_per_semester(self):
        for semester in self.semester_info:
            validate_semester=semester.semester_validance()
            color="green" if validate_semester else "red"
            termcolor.cprint(semester.id+":"+str(len(semester.courses))+" courses",color)
    
    def export_program_by_lecturer(self,lecturer_identifier):
        lecturer=self.lecturers[self.lecturers.index(lecturer_identifier)]
        print(f'{lecturer.name}({lecturer.rank.name})')
        lecturer.info()
        
            
