import openpyxl
import os
from prettytable import PrettyTable
from pathlib import Path

def main():
    path_to_xlsxs=os.path.join('','CBT_Source')
    files=[file for file in os.listdir(path_to_xlsxs) if file.endswith('xlsx')]
    print(files)
    for file in files:
        active_path=os.path.join(path_to_xlsxs,file)
        wb_obj=openpyxl.load_workbook(active_path)
        sheet_obj=wb_obj.active
        rows=sheet_obj.max_row
        columns=sheet_obj.max_column
        table=PrettyTable()
        for i in range(1,rows+1):
            row=list()
            for j in range(1,columns+1):
                cell_obj=sheet_obj.cell(row=i,column=j)
                row.append(cell_obj.value)
            if i==1:
                table.field_names=row
            else:
                table.add_row(row)
        print(table,'\n\n')




if __name__=='__main__':
    main()